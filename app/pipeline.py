from __future__ import annotations

import json
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .entity_resolution import EntityResolutionConfig, resolve_employees
from .location_dictionary import enrich_location_dictionary_in_place, implicit_location_dictionary
from .normalize import events_to_frame
from .report_discovery import ReportType, discover_reports
from .report_parsers import parse_absence_details, parse_export_travel, parse_hr_attendance, parse_remote_working_request, parse_transport_pdf
from .rules_engine import RulesConfig, infer_employee_locations_with_trace
from .utils import parse_dt

logger = logging.getLogger(__name__)

PARSERS_BY_TYPE = {
    "remote_working_request": lambda report, cfg, source_label, priority: parse_remote_working_request(
        report.path,
        asof_tz=cfg.asof_timezone,
        source_label=source_label,
        source_priority=priority,
    ),
    "hr_attendance": lambda report, cfg, source_label, priority: parse_hr_attendance(
        report.path,
        asof_tz=cfg.asof_timezone,
        source_label=source_label,
        source_priority=priority,
    ),
    "absence_details": lambda report, cfg, source_label, priority: parse_absence_details(
        report.path,
        asof_tz=cfg.asof_timezone,
        source_label=source_label,
        source_priority=priority,
    ),
    "export_travel": lambda report, cfg, source_label, priority: parse_export_travel(
        report.path,
        asof_tz=cfg.asof_timezone,
        source_label=source_label,
        source_priority=priority,
    ),
    "transport_arrival": lambda report, cfg, source_label, priority: parse_transport_pdf(
        report.path,
        side="arrival",
        asof_tz=cfg.asof_timezone,
        source_label=source_label,
        source_priority=priority,
    ),
    "transport_departure": lambda report, cfg, source_label, priority: parse_transport_pdf(
        report.path,
        side="departure",
        asof_tz=cfg.asof_timezone,
        source_label=source_label,
        source_priority=priority,
    ),
}


@dataclass(frozen=True)
class MVPConfig:
    asof_timezone: str
    source_priorities_by_type: dict[ReportType, int]
    fuzzy_enabled: bool
    fuzzy_threshold: int
    rules: RulesConfig
    outputs_dir: str
    output_csv: str
    output_json: str
    output_excel: str


def load_config(path: str) -> MVPConfig:
    config_path = Path(path).resolve()
    with open(path, "r", encoding="utf-8") as f:
        raw: dict[str, Any] = json.load(f)

    rules_raw = raw.get("rules", {})
    rules = RulesConfig(
        event_type_priority=list(rules_raw.get("event_type_priority", [])),
        office_checkin_valid_hours=int(rules_raw.get("office_checkin_valid_hours", 16)),
        working_format_default_location=str(rules_raw.get("working_format_default_location", "REMOTE")),
    )

    out_raw = raw.get("outputs", {})

    source_priorities_raw = raw.get("source_priorities_by_type", {})
    source_priorities_by_type: dict[ReportType, int] = {}
    for k, v in source_priorities_raw.items():
        source_priorities_by_type[k] = int(v)

    return MVPConfig(
        asof_timezone=str(raw.get("asof_timezone", "UTC")),
        source_priorities_by_type=source_priorities_by_type,
        fuzzy_enabled=bool(raw.get("fuzzy_matching", {}).get("enabled", True)),
        fuzzy_threshold=int(raw.get("fuzzy_matching", {}).get("threshold", 92)),
        rules=rules,
        outputs_dir=str(out_raw.get("dir", "out")),
        output_csv=str(out_raw.get("csv", "employee_locations.csv")),
        output_json=str(out_raw.get("json", "employee_locations.json")),
        output_excel=str(out_raw.get("excel", "employee_locations.xlsx")),
    )


def _autosize_worksheet(ws) -> None:
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 40)


def _style_worksheet(ws, *, freeze_cell: str = "A2") -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions
    _autosize_worksheet(ws)


def _build_presentable_locations(locations: pd.DataFrame) -> pd.DataFrame:
    presentable = locations.copy()
    presentable = presentable[
        [
            "resolved_employee_id",
            "resolved_name",
            "location",
            "chosen_start_ts",
            "chosen_event_type",
            "chosen_source",
        ]
    ].rename(
        columns={
            "resolved_employee_id": "Employee ID",
            "resolved_name": "Employee Name",
            "location": "Location",
            "chosen_start_ts": "Recorded at",
            "chosen_event_type": "Based on",
            "chosen_source": "Report file",
        }
    )
    return presentable


def presentable_locations_df(locations: pd.DataFrame) -> pd.DataFrame:
    """Same columns as the Employee Locations Excel sheet (human-readable headers)."""
    return _build_presentable_locations(locations)


def _write_polished_excel(
    *,
    excel_path: Path,
    locations: pd.DataFrame,
    resolved_events: pd.DataFrame,
    decision_trace: pd.DataFrame | None,
    asof: datetime,
) -> Path:
    summary = _build_presentable_locations(locations)
    report_info = pd.DataFrame(
        [
            {"Field": "Report generated for", "Value": asof.isoformat()},
            {"Field": "Recorded at", "Value": "Timestamp of the event used to determine the employee's location"},
            {"Field": "Based on", "Value": "Type of evidence that won in the rules engine (vacation, office check-in, travel, etc.)"},
        ]
    )

    evidence = resolved_events.copy()
    evidence = evidence.sort_values(
        by=["resolved_name", "employee_key", "start_ts"],
        ascending=[True, True, False],
    )

    source_summary = (
        resolved_events.groupby(["source", "event_type"], as_index=False)
        .size()
        .rename(columns={"size": "rows"})
        .sort_values(by=["source", "event_type"])
    )

    # Excel does not support timezone-aware datetimes.
    dfs_for_tz = [summary, evidence]
    if decision_trace is not None:
        dfs_for_tz.append(decision_trace)
    for df in dfs_for_tz:
        for col in df.columns:
            if pd.api.types.is_datetime64tz_dtype(df[col]):
                df[col] = df[col].dt.tz_localize(None)

    candidate_paths = [excel_path]
    fallback_name = f"{excel_path.stem}_{asof.strftime('%Y%m%d_%H%M%S')}{excel_path.suffix}"
    candidate_paths.append(excel_path.with_name(fallback_name))

    last_error: Exception | None = None
    for candidate in candidate_paths:
        try:
            with pd.ExcelWriter(candidate, engine="openpyxl") as writer:
                summary.to_excel(writer, sheet_name="Employee Locations", index=False)
                report_info.to_excel(writer, sheet_name="Report Info", index=False)
                evidence.to_excel(writer, sheet_name="Evidence", index=False)
                source_summary.to_excel(writer, sheet_name="Source Summary", index=False)
                if decision_trace is not None:
                    decision_trace.to_excel(writer, sheet_name="Decision Trace", index=False)

                wb = writer.book
                ws_locations = wb["Employee Locations"]
                ws_info = wb["Report Info"]
                ws_evidence = wb["Evidence"]
                ws_sources = wb["Source Summary"]
                ws_trace = wb["Decision Trace"] if decision_trace is not None else None

                _style_worksheet(ws_locations)
                _style_worksheet(ws_info)
                _style_worksheet(ws_evidence)
                _style_worksheet(ws_sources)
                if ws_trace is not None:
                    _style_worksheet(ws_trace)

                datetime_columns = {
                    "Employee Locations": ["E"],
                    "Evidence": ["E", "F"],
                    "Decision Trace": ["M", "N"],
                }
                for sheet_name, cols in datetime_columns.items():
                    if sheet_name not in wb.sheetnames:
                        continue
                    ws = wb[sheet_name]
                    for col in cols:
                        for cell in ws[col][1:]:
                            cell.number_format = "yyyy-mm-dd hh:mm"

                ws_locations.sheet_view.zoomScale = 90
                ws_info.sheet_view.zoomScale = 95
                ws_evidence.sheet_view.zoomScale = 85
                ws_sources.sheet_view.zoomScale = 90
            return candidate
        except PermissionError as exc:
            last_error = exc
            continue

    if last_error is not None:
        raise last_error
    return excel_path


def _safe_write_table_outputs(
    *,
    locations: pd.DataFrame,
    csv_path: Path,
    json_path: Path,
) -> None:
    presentable = _build_presentable_locations(locations)
    try:
        presentable.to_csv(csv_path, index=False)
    except PermissionError:
        pass

    try:
        presentable.to_json(json_path, orient="records", date_format="iso")
    except PermissionError:
        pass


def run_pipeline(
    *,
    config_path: str,
    asof: datetime,
    input_dir: str = "in",
    output_dir: str | None = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    cfg = load_config(config_path)

    location_dict = implicit_location_dictionary()

    logger.info("Run started. asof=%s input_dir=%s config=%s", asof.isoformat(), input_dir, config_path)

    all_events = []
    reports = discover_reports(input_dir)
    logger.info("Discovered %d report(s).", len(reports))
    parse_failures: list[str] = []
    for report in reports:
        priority = cfg.source_priorities_by_type.get(report.report_type, 50)
        source_label = report.source_label or report.path.name
        logger.info(
            "Parsing report. type=%s path=%s source=%s priority=%s",
            report.report_type,
            str(report.path),
            source_label,
            priority,
        )
        try:
            parser = PARSERS_BY_TYPE.get(report.report_type)
            if parser is None:
                raise ValueError(f"No parser registered for report type: {report.report_type}")
            events = parser(report, cfg, source_label, priority)
        except Exception as exc:
            msg = f"{report.report_type}::{report.path.name} - {exc}"
            parse_failures.append(msg)
            logger.exception("Failed to parse report: %s", msg)
            continue

        logger.info("Parsed %d event(s) from %s.", len(events), source_label)
        all_events.extend(events)

    events_df = events_to_frame(all_events)
    if events_df.empty:
        if parse_failures:
            raise RuntimeError("No events loaded from sources. Parse failures:\n" + "\n".join(parse_failures))
        raise RuntimeError("No events loaded from sources.")

    logger.info("Total events loaded: %d", len(events_df.index))

    # Make sure any previously unseen codes (e.g. airport/site codes) are at least visible as themselves.
    try:
        added = enrich_location_dictionary_in_place(location_dict, raw_values=events_df["location_raw"])
        if added:
            logger.info("Location dictionary enriched with %d new code(s) from inputs.", added)
    except Exception:
        logger.exception("Failed to enrich location dictionary from inputs.")

    resolved_events = resolve_employees(
        events_df,
        EntityResolutionConfig(
            fuzzy_enabled=cfg.fuzzy_enabled,
            fuzzy_threshold=cfg.fuzzy_threshold,
        ),
    )
    logger.info(
        "Entity resolution complete. unique_employees=%d",
        resolved_events["employee_key"].nunique(),
    )

    locations, trace = infer_employee_locations_with_trace(
        events=resolved_events,
        asof=asof,
        rules=cfg.rules,
        location_dict=location_dict,
    )
    logger.info("Rules engine complete. locations_rows=%d", len(locations.index))

    # Write outputs
    out_dir = Path(output_dir or cfg.outputs_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    csv_path = out_dir / cfg.output_csv
    json_path = out_dir / cfg.output_json
    excel_path = out_dir / cfg.output_excel

    _safe_write_table_outputs(
        locations=locations,
        csv_path=csv_path,
        json_path=json_path,
    )
    _write_polished_excel(
        excel_path=excel_path,
        locations=locations,
        resolved_events=resolved_events,
        decision_trace=trace,
        asof=asof,
    )

    logger.info("Outputs written. out_dir=%s", str(out_dir))
    return resolved_events, locations


def parse_asof(asof_str: str, tz: str = "UTC") -> datetime:
    return parse_dt(asof_str, tz=tz)

